from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
PROFIT_FILL = PatternFill("solid", fgColor="DCFCE7")
LOSS_FILL = PatternFill("solid", fgColor="FEE2E2")
TITLE_FILL = PatternFill("solid", fgColor="0F172A")


class ExcelReportGenerator:
    def generate(self, result: dict[str, Any], output_path: str | Path) -> Path:
        workbook = Workbook()
        workbook.remove(workbook.active)

        strategy_results = result.get("strategy_results", [])
        ranking_rows = result.get("ranking", {}).get("ranking", [])
        best = result.get("ranking", {}).get("best_overall_strategy") or {}
        all_trades = [trade for strategy in strategy_results for trade in strategy.get("trades", [])]
        monthly_rows = self._flatten(strategy_results, "monthly_returns", "month")
        yearly_rows = self._flatten(strategy_results, "yearly_returns", "year")
        drawdown_rows = self._curve_rows(strategy_results, "drawdown_curve")
        equity_rows = self._curve_rows(strategy_results, "equity_curve")
        expiry_rows = self._flatten(strategy_results, "expiry_analysis", "expiry")
        risk_rows = [self._risk_row(item) for item in strategy_results]
        explain_rows = self._explanations(strategy_results)
        best_worst_rows = self._best_worst(strategy_results)
        index_comparison = self._index_comparison(strategy_results)

        self._dashboard_sheet(workbook.create_sheet("Dashboard"), result, best)
        self._table_sheet(workbook.create_sheet("Strategy Ranking"), ranking_rows)
        self._table_sheet(workbook.create_sheet("Trade Book"), all_trades)
        self._table_sheet(workbook.create_sheet("Monthly P&L"), monthly_rows)
        self._table_sheet(workbook.create_sheet("Yearly P&L from 2006"), yearly_rows)
        self._table_sheet(workbook.create_sheet("Drawdown"), drawdown_rows)
        self._table_sheet(workbook.create_sheet("Equity Curve"), equity_rows)
        self._table_sheet(workbook.create_sheet("Index Comparison"), index_comparison)
        self._table_sheet(workbook.create_sheet("Expiry Day Analysis"), expiry_rows)
        self._table_sheet(workbook.create_sheet("Best Worst Trades"), best_worst_rows)
        self._table_sheet(workbook.create_sheet("Risk Metrics"), risk_rows)
        self._table_sheet(workbook.create_sheet("AI Signal Explanation"), explain_rows)

        self._add_charts(workbook)

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
        return output

    def _dashboard_sheet(self, sheet, result: dict[str, Any], best: dict[str, Any]) -> None:
        sheet["A1"] = "Antigravity Backtesting Dashboard"
        sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        sheet["A1"].fill = TITLE_FILL
        sheet["A3"] = "Job ID"
        sheet["B3"] = result.get("job_id", "")
        sheet["A4"] = "Index"
        sheet["B4"] = result.get("request", {}).get("index", "")
        sheet["A5"] = "Strategy"
        sheet["B5"] = result.get("request", {}).get("strategy", "")
        sheet["A6"] = "Date Range"
        req = result.get("request", {})
        sheet["B6"] = f"{req.get('start_date', '')} -> {req.get('end_date', '')}"
        sheet["A8"] = "Best Overall Strategy"
        sheet["B8"] = best.get("strategy", "N/A")
        sheet["A9"] = "AI Score"
        sheet["B9"] = best.get("score", 0)
        sheet["A10"] = "Net P&L"
        sheet["B10"] = best.get("net_pnl", 0)
        self._format_header_row(sheet, 3, 2)
        self._auto_width(sheet)

    def _table_sheet(self, sheet, rows: list[dict[str, Any]]) -> None:
        if not rows:
            sheet.append(["message"])
            sheet.append(["No data"])
            self._format_header_row(sheet, 1, 1)
            self._auto_width(sheet)
            return
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            values = []
            for header in headers:
                value = row.get(header)
                if isinstance(value, list):
                    values.append(", ".join(str(item) for item in value))
                elif isinstance(value, dict):
                    values.append(str(value))
                else:
                    values.append(value)
            sheet.append(values)
        self._format_header_row(sheet, 1, len(headers))
        self._style_profit_loss(sheet)
        self._auto_width(sheet)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    def _format_header_row(self, sheet, row_number: int, columns: int) -> None:
        for idx in range(1, columns + 1):
            cell = sheet.cell(row=row_number, column=idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = HEADER_FILL

    def _style_profit_loss(self, sheet) -> None:
        headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}
        for header in ["net_pnl", "gross_pnl", "drawdown", "equity"]:
            if header not in headers:
                continue
            col = headers[header]
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.fill = PROFIT_FILL if cell.value >= 0 else LOSS_FILL

    def _auto_width(self, sheet) -> None:
        for column in sheet.columns:
            length = 0
            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                length = max(length, len(value))
            sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(length + 2, 12), 42)

    def _flatten(self, strategy_results: list[dict[str, Any]], metric_key: str, dimension: str) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for strategy in strategy_results:
            for item in strategy.get("metrics", {}).get(metric_key, []):
                rows.append({"strategy": strategy["strategy"], "index": strategy["index"], dimension: item.get(dimension), **item})
        return rows

    def _curve_rows(self, strategy_results: list[dict[str, Any]], metric_key: str) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for strategy in strategy_results:
            for item in strategy.get("metrics", {}).get(metric_key, []):
                rows.append({"strategy": strategy["strategy"], "index": strategy["index"], **item})
        return rows

    def _risk_row(self, item: dict[str, Any]) -> dict[str, Any]:
        metrics = item.get("metrics", {})
        return {
            "strategy": item["strategy"],
            "index": item["index"],
            "win_rate": metrics.get("win_rate"),
            "profit_factor": metrics.get("profit_factor"),
            "max_drawdown_pct": metrics.get("max_drawdown_pct"),
            "sharpe_ratio": metrics.get("sharpe_ratio"),
            "sortino_ratio": metrics.get("sortino_ratio"),
            "stability": metrics.get("stability"),
            "strategy_score": metrics.get("strategy_score"),
        }

    def _explanations(self, strategy_results: list[dict[str, Any]]) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for strategy in strategy_results:
            for trade in strategy.get("trades", []):
                rows.append(
                    {
                        "strategy": strategy["strategy"],
                        "index": strategy["index"],
                        "entry_time": trade.get("entry_time"),
                        "signal": trade.get("signal"),
                        "confidence": trade.get("confidence"),
                        "reasons": " | ".join(trade.get("reasons", [])),
                    }
                )
        return rows

    def _best_worst(self, strategy_results: list[dict[str, Any]]) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for strategy in strategy_results:
            metrics = strategy.get("metrics", {})
            if metrics.get("best_trade"):
                rows.append({"type": "BEST", "strategy": strategy["strategy"], "index": strategy["index"], **metrics["best_trade"]})
            if metrics.get("worst_trade"):
                rows.append({"type": "WORST", "strategy": strategy["strategy"], "index": strategy["index"], **metrics["worst_trade"]})
        return rows

    def _index_comparison(self, strategy_results: list[dict[str, Any]]) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for strategy in strategy_results:
            metrics = strategy["metrics"]
            rows.append(
                {
                    "index": strategy["index"],
                    "strategy": strategy["strategy"],
                    "net_pnl": metrics.get("net_pnl"),
                    "win_rate": metrics.get("win_rate"),
                    "score": metrics.get("strategy_score"),
                }
            )
        return rows

    def _add_charts(self, workbook: Workbook) -> None:
        dashboard = workbook["Dashboard"]

        if "Equity Curve" in workbook and workbook["Equity Curve"].max_row > 1:
            sheet = workbook["Equity Curve"]
            chart = LineChart()
            chart.title = "Equity Curve"
            chart.y_axis.title = "Equity"
            data = Reference(sheet, min_col=4, max_col=4, min_row=1, max_row=sheet.max_row)
            cats = Reference(sheet, min_col=3, max_col=3, min_row=2, max_row=sheet.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            dashboard.add_chart(chart, "D3")

        if "Drawdown" in workbook and workbook["Drawdown"].max_row > 1:
            sheet = workbook["Drawdown"]
            chart = LineChart()
            chart.title = "Drawdown"
            chart.y_axis.title = "Drawdown"
            data = Reference(sheet, min_col=4, max_col=4, min_row=1, max_row=sheet.max_row)
            cats = Reference(sheet, min_col=3, max_col=3, min_row=2, max_row=sheet.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            dashboard.add_chart(chart, "D18")

        if "Monthly P&L" in workbook and workbook["Monthly P&L"].max_row > 1:
            sheet = workbook["Monthly P&L"]
            chart = BarChart()
            chart.title = "Monthly P&L"
            chart.y_axis.title = "Net P&L"
            data = Reference(sheet, min_col=5, max_col=5, min_row=1, max_row=sheet.max_row)
            cats = Reference(sheet, min_col=3, max_col=3, min_row=2, max_row=sheet.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            dashboard.add_chart(chart, "L3")

        if "Strategy Ranking" in workbook and workbook["Strategy Ranking"].max_row > 1:
            sheet = workbook["Strategy Ranking"]
            chart = BarChart()
            chart.title = "Strategy Comparison"
            chart.y_axis.title = "AI Score"
            data = Reference(sheet, min_col=4, max_col=4, min_row=1, max_row=sheet.max_row)
            cats = Reference(sheet, min_col=1, max_col=1, min_row=2, max_row=sheet.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            dashboard.add_chart(chart, "L18")
