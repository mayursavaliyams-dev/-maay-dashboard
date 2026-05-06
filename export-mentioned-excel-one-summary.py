from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE_FILE = Path(
    "exports/daily-all-indices-6trades-equity-500000-position-2500-nolimit-ladder-charges-0.09-net-pnl-after-charges.xlsx"
)
OUT_FILE = Path("exports/one-excel-one-summary-from-net-pnl-after-charges.xlsx")
INDEX_SHEETS = ["SENSEX", "NIFTY", "BANKNIFTY"]


def number(series):
    return pd.to_numeric(series, errors="coerce")


def summarize_sheet(source, sheet_name):
    df = pd.read_excel(source, sheet_name=sheet_name)
    if df.empty:
        return None

    net = number(df.get("NetPnl", pd.Series(dtype=float))).fillna(0)
    gross = number(df.get("GrossPnl", pd.Series(dtype=float))).fillna(0)
    charges = number(df.get("TotalCharges", pd.Series(dtype=float))).fillna(0)
    result = df.get("Result", pd.Series(dtype=str)).astype(str).str.upper()
    raw_result = df.get("RawResult", pd.Series(dtype=str)).astype(str).str.upper()
    trade_date = pd.to_datetime(df.get("Date"), errors="coerce")
    opening = number(df.get("StartEquity", pd.Series(dtype=float))).dropna()
    closing = number(df.get("EndEquity", pd.Series(dtype=float))).dropna()
    drawdown = number(df.get("DrawdownPct", pd.Series(dtype=float))).dropna()
    multiplier = number(df.get("Multiplier", pd.Series(dtype=float))).dropna()
    position = number(df.get("PositionSize", pd.Series(dtype=float))).dropna()

    wins = int((result == "WIN").sum())
    losses = int((result == "LOSS").sum())
    raw_wins = int((raw_result == "WIN").sum())
    raw_losses = int((raw_result == "LOSS").sum())
    total_trades = int(len(df))
    gross_profit = float(gross[gross > 0].sum())
    gross_loss = float(gross[gross < 0].sum())
    net_profit = float(net[net > 0].sum())
    net_loss = float(net[net < 0].sum())
    total_charges = float(charges.sum())
    net_pnl = float(net.sum())
    opening_capital = float(opening.iloc[0]) if not opening.empty else 0.0
    closing_capital = float(closing.iloc[-1]) if not closing.empty else opening_capital + net_pnl
    average_profit = float(net[net > 0].mean()) if (net > 0).any() else 0.0
    average_loss = float(net[net < 0].mean()) if (net < 0).any() else 0.0
    profit_factor = gross_profit / abs(gross_loss) if gross_loss else 0.0
    payoff_ratio = average_profit / abs(average_loss) if average_loss else 0.0

    return {
        "Instrument": sheet_name,
        "Start Date": trade_date.min().date().isoformat() if trade_date.notna().any() else "",
        "End Date": trade_date.max().date().isoformat() if trade_date.notna().any() else "",
        "Trading Days": int(trade_date.dt.date.nunique()) if trade_date.notna().any() else 0,
        "Total Trades": total_trades,
        "Winning Trades": wins,
        "Losing Trades": losses,
        "Win Rate %": wins / total_trades * 100 if total_trades else 0.0,
        "Raw Winning Trades": raw_wins,
        "Raw Losing Trades": raw_losses,
        "Raw Win Rate %": raw_wins / total_trades * 100 if total_trades else 0.0,
        "Opening Capital ₹": opening_capital,
        "Closing Capital ₹": closing_capital,
        "Gross Profit ₹": gross_profit,
        "Gross Loss ₹": gross_loss,
        "Gross P&L ₹": gross_profit + gross_loss,
        "Net Profit ₹": net_profit,
        "Net Loss ₹": net_loss,
        "Net P&L ₹": net_pnl,
        "Total Charges ₹": total_charges,
        "Net P&L After Charges ₹": net_pnl,
        "Return %": ((closing_capital - opening_capital) / opening_capital * 100) if opening_capital else 0.0,
        "Average Profit ₹": average_profit,
        "Average Loss ₹": average_loss,
        "Profit Factor": profit_factor,
        "Payoff Ratio": payoff_ratio,
        "Expectancy / Trade ₹": net_pnl / total_trades if total_trades else 0.0,
        "Max Drawdown %": float(drawdown.max()) if not drawdown.empty else 0.0,
        "Average Multiplier": float(multiplier.mean()) if not multiplier.empty else 0.0,
        "Max Multiplier": float(multiplier.max()) if not multiplier.empty else 0.0,
        "Min Position ₹": float(position.min()) if not position.empty else 2500.0,
        "Max Position": "NO LIMIT",
        "Position Rule": "start 2500, no max position",
        "Max Trades / Day": 6,
        "Charges Per Order %": 0.09,
        "Source File": SOURCE_FILE.name,
    }


def whole_total(rows):
    df = pd.DataFrame(rows)
    trades = float(df["Total Trades"].sum())
    wins = float(df["Winning Trades"].sum())
    losses = float(df["Losing Trades"].sum())
    gross_profit = float(df["Gross Profit ₹"].sum())
    gross_loss = float(df["Gross Loss ₹"].sum())
    net_profit = float(df["Net Profit ₹"].sum())
    net_loss = float(df["Net Loss ₹"].sum())
    net_pnl = float(df["Net P&L ₹"].sum())
    opening = float(df["Opening Capital ₹"].sum())
    closing = float(df["Closing Capital ₹"].sum())
    total_charges = float(df["Total Charges ₹"].sum())

    return {
        "Instrument": "WHOLE TOTAL",
        "Start Date": df["Start Date"].min(),
        "End Date": df["End Date"].max(),
        "Trading Days": "",
        "Total Trades": trades,
        "Winning Trades": wins,
        "Losing Trades": losses,
        "Win Rate %": wins / trades * 100 if trades else 0.0,
        "Raw Winning Trades": float(df["Raw Winning Trades"].sum()),
        "Raw Losing Trades": float(df["Raw Losing Trades"].sum()),
        "Raw Win Rate %": float(df["Raw Winning Trades"].sum()) / trades * 100 if trades else 0.0,
        "Opening Capital ₹": opening,
        "Closing Capital ₹": closing,
        "Gross Profit ₹": gross_profit,
        "Gross Loss ₹": gross_loss,
        "Gross P&L ₹": gross_profit + gross_loss,
        "Net Profit ₹": net_profit,
        "Net Loss ₹": net_loss,
        "Net P&L ₹": net_pnl,
        "Total Charges ₹": total_charges,
        "Net P&L After Charges ₹": net_pnl,
        "Return %": ((closing - opening) / opening * 100) if opening else 0.0,
        "Average Profit ₹": "",
        "Average Loss ₹": "",
        "Profit Factor": gross_profit / abs(gross_loss) if gross_loss else 0.0,
        "Payoff Ratio": "",
        "Expectancy / Trade ₹": net_pnl / trades if trades else 0.0,
        "Max Drawdown %": float(df["Max Drawdown %"].max()),
        "Average Multiplier": float(df["Average Multiplier"].mean()),
        "Max Multiplier": float(df["Max Multiplier"].max()),
        "Min Position ₹": 2500,
        "Max Position": "NO LIMIT",
        "Position Rule": "start 2500, no max position",
        "Max Trades / Day": 6,
        "Charges Per Order %": 0.09,
        "Source File": SOURCE_FILE.name,
    }


def format_workbook(path):
    wb = load_workbook(path)
    ws = wb["One Summary"]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = total_fill

    money_cols = [cell.column for cell in ws[1] if "₹" in str(cell.value)]
    pct_cols = [cell.column for cell in ws[1] if "%" in str(cell.value)]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in money_cols:
                cell.number_format = '₹#,##0.00'
            elif cell.column in pct_cols:
                cell.number_format = '0.00'
            elif isinstance(cell.value, float):
                cell.number_format = '#,##0.00'

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in ws[letter])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 34)
    ws.freeze_panes = "A2"
    wb.save(path)


def main():
    OUT_FILE.parent.mkdir(exist_ok=True)
    rows = [summarize_sheet(SOURCE_FILE, sheet) for sheet in INDEX_SHEETS]
    rows = [row for row in rows if row]
    output = [whole_total(rows), *rows]

    with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as writer:
        pd.DataFrame(output).to_excel(writer, sheet_name="One Summary", index=False)

    format_workbook(OUT_FILE)
    print(OUT_FILE.resolve())


if __name__ == "__main__":
    main()
