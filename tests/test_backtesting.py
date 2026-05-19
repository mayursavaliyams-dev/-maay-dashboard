from __future__ import annotations

from pathlib import Path

import pandas as pd
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.analytics import GammaBlastAnalyzer, GreeksMatrixAnalyzer, OptionsAnalyticsAnalyzer
from backend.main import app
from backend.backtesting.data_loader import HistoricalDataEngine, normalize_timeframe_rule
from backend.backtesting.engine import BacktestEngine, BacktestRunRequest
from backend.backtesting.metrics import calculate_metrics
from backend.backtesting.report_excel import ExcelReportGenerator
from backend.backtesting.strategies.premium_breakout import PremiumBreakoutStrategy


def build_sample_csv(path: Path) -> Path:
    rows = []
    expiry = "2026-05-07"
    strike_pack = [24400, 24500, 24600]
    day_specs = [("2026-05-01", 24400, 1), ("2026-05-02", 24550, -1)]
    for day, base_spot, direction in day_specs:
        times = pd.date_range(f"{day} 09:15:00", periods=14, freq="5min")
        for idx, dt in enumerate(times):
            spot = base_spot + direction * idx * 22
            rows.append(
                {
                    "datetime": dt,
                    "index": "NIFTY",
                    "record_type": "SPOT",
                    "expiry": "",
                    "strike": "",
                    "option_type": "",
                    "open": spot - 5,
                    "high": spot + 8,
                    "low": spot - 9,
                    "close": spot,
                    "ltp": spot,
                    "volume": 0,
                    "oi": 0,
                    "iv": "",
                    "delta": "",
                    "gamma": "",
                    "theta": "",
                    "vega": "",
                    "spot_open": spot - 5,
                    "spot_high": spot + 8,
                    "spot_low": spot - 9,
                    "spot_close": spot,
                    "futures_close": spot + 4,
                }
            )
            for strike in strike_pack:
                distance = strike - spot
                ce_price = max(18, 115 + direction * idx * 12 - max(distance, 0) * 0.08 + max(-distance, 0) * 0.02)
                pe_price = max(16, 108 - direction * idx * 11 + max(distance, 0) * 0.03 + max(-distance, 0) * 0.09)
                for option_type, price, delta in [("CE", ce_price, 0.42), ("PE", pe_price, -0.43)]:
                    rows.append(
                        {
                            "datetime": dt,
                            "index": "NIFTY",
                            "record_type": "OPTION",
                            "expiry": expiry,
                            "strike": strike,
                            "option_type": option_type,
                            "open": round(price * 0.98, 2),
                            "high": round(price * 1.05, 2),
                            "low": round(price * 0.96, 2),
                            "close": round(price, 2),
                            "ltp": round(price, 2),
                            "volume": 15000 + idx * 1100 + (400 if strike == 24500 else 0),
                            "oi": 200000 + idx * 3500 + (1200 if option_type == "CE" else 900),
                            "iv": 13.5 + idx * 0.2,
                            "delta": delta,
                            "gamma": 0.012 + idx * 0.0004,
                            "theta": -2.4 - idx * 0.05,
                            "vega": 8.0 + idx * 0.1,
                            "spot_open": spot - 5,
                            "spot_high": spot + 8,
                            "spot_low": spot - 9,
                            "spot_close": spot,
                            "futures_close": spot + 4,
                        }
                    )
    pd.DataFrame(rows).to_csv(path, index=False)
    return path


def test_data_loading(tmp_path: Path):
    csv_path = build_sample_csv(tmp_path / "sample.csv")
    loader = HistoricalDataEngine()
    dataset = loader.load_csv(csv_path, timeframe="5m", index_filter="NIFTY")
    assert not dataset.options.empty
    assert "NIFTY" in dataset.metadata["indices"]
    assert {"datetime", "spot_close", "volume_ratio", "premium_jump_pct"}.issubset(dataset.options.columns)


def test_strategy_signal_generation(tmp_path: Path):
    csv_path = build_sample_csv(tmp_path / "sample.csv")
    dataset = HistoricalDataEngine().load_csv(csv_path, timeframe="5m", index_filter="NIFTY")
    strategy = PremiumBreakoutStrategy()
    signals = strategy.generate_signals(dataset, context=strategy_context())
    assert signals
    assert {signal.signal for signal in signals}.issubset({"BUY_CALL", "BUY_PUT"})


def test_custom_timeframe_normalization_and_resampling(tmp_path: Path):
    csv_path = build_sample_csv(tmp_path / "sample.csv")
    dataset = HistoricalDataEngine().load_csv(csv_path, timeframe="10m", index_filter="NIFTY")
    assert normalize_timeframe_rule("10m") == "10min"
    assert normalize_timeframe_rule("60m") == "60min"
    assert not dataset.options.empty
    assert dataset.metadata["timeframe"] == "10m"


def test_backtest_run(tmp_path: Path):
    csv_path = build_sample_csv(tmp_path / "sample.csv")
    engine = BacktestEngine()
    request = BacktestRunRequest(
        index="NIFTY",
        strategy="combined_ai",
        start_date="2026-05-01",
        end_date="2026-05-02",
        capital=200000,
        lot_size=50,
        timeframe="5m",
    )
    result = engine.run(request, dataset_path=csv_path, job_id="unit_job", report_dir=tmp_path)
    assert result["strategy_results"]
    assert Path(result["report_path"]).exists()
    assert result["ranking"]["ranking"]


def test_metrics_calculation():
    trades = [
        {"entry_time": "2026-05-01T09:20:00", "exit_time": "2026-05-01T09:45:00", "strategy": "A", "index": "NIFTY", "signal": "BUY_CALL", "strike": 24500, "return_pct": 12.5, "net_pnl": 2500, "gross_pnl": 2540, "exit_reason": "TARGET", "market_regime": "TRENDING", "volatility_regime": "HIGH_VOL"},
        {"entry_time": "2026-05-02T10:00:00", "exit_time": "2026-05-02T10:25:00", "strategy": "A", "index": "NIFTY", "signal": "BUY_PUT", "strike": 24500, "return_pct": -5.2, "net_pnl": -1200, "gross_pnl": -1160, "exit_reason": "STOP_LOSS", "market_regime": "SIDEWAYS", "volatility_regime": "LOW_VOL"},
    ]
    metrics = calculate_metrics(trades, initial_capital=100000, start_date="2026-05-01", end_date="2026-05-02")
    assert metrics["total_trades"] == 2
    assert metrics["winning_trades"] == 1
    assert metrics["net_pnl"] == 1300.0
    assert metrics["best_trade"]["exit_reason"] == "TARGET"


def test_excel_report_generation(tmp_path: Path):
    csv_path = build_sample_csv(tmp_path / "sample.csv")
    engine = BacktestEngine()
    request = BacktestRunRequest(index="NIFTY", strategy="premium_breakout", start_date="2026-05-01", end_date="2026-05-02", capital=200000, lot_size=50, timeframe="5m")
    result = engine.run(request, dataset_path=csv_path, job_id="excel_job", report_dir=tmp_path)
    report_path = Path(result["report_path"])
    workbook = load_workbook(report_path)
    assert "Dashboard" in workbook.sheetnames
    assert "Strategy Ranking" in workbook.sheetnames
    assert "Trade Book" in workbook.sheetnames
    assert "Best Worst Trades" in workbook.sheetnames


def test_gamma_blast_analyzer_ranges():
    analyzer = GammaBlastAnalyzer()
    result = analyzer.analyze(spot_price=75000, index="SENSEX", iv_pct=15.0, time_to_expiry_days=1.0)
    assert result["blastLevel"] in {"LOW", "MODERATE", "HIGH", "EXTREME", "NUCLEAR"}
    assert result["possibleRanges"]["atmStrike"] % 100 == 0
    assert len(result["possibleRanges"]["bands"]) == 3
    assert result["possibleRanges"]["bands"][0]["low"] < result["possibleRanges"]["bands"][0]["high"]


def test_gamma_blast_endpoint():
    client = TestClient(app)
    response = client.get("/api/options/gamma-blast", params={"spot_price": 75000, "index": "SENSEX", "iv_pct": 15, "time_to_expiry_days": 1})
    assert response.status_code == 200
    payload = response.json()
    assert payload["index"] == "SENSEX"
    assert payload["possibleRanges"]["bands"][1]["label"] == "Expected"
    assert "metrics" in payload and "ladder" in payload


def test_greeks_matrix_analyzer():
    analyzer = GreeksMatrixAnalyzer()
    result = analyzer.analyze(spot_price=75000, index="SENSEX", iv_pct=15.0, time_to_expiry_days=1.0)
    assert result["inst"] == "SENSEX"
    assert result["interval"] == 100
    assert len(result["rows"]) == 5
    assert result["summary"]["atmStrike"] == result["atm"]
    assert result["rows"][2]["isATM"] is True


def test_greeks_matrix_endpoint():
    client = TestClient(app)
    response = client.get("/api/options/greeks-matrix", params={"spot_price": 75000, "index": "SENSEX", "iv_pct": 15, "time_to_expiry_days": 1})
    assert response.status_code == 200
    payload = response.json()
    assert payload["inst"] == "SENSEX"
    assert payload["summary"]["atmStrike"] == payload["atm"]
    assert len(payload["rows"]) == 5


def test_options_analytics_analyzer_complete_payload():
    analyzer = OptionsAnalyticsAnalyzer()
    result = analyzer.analyze_complete(spot_price=24530, index="NIFTY", iv_pct=14.0, time_to_expiry_days=2.0)
    assert result["index"] == "NIFTY"
    assert result["atmStrike"] % 50 == 0
    assert len(result["optionChain"]) == 21
    assert result["pcr"]["interpretation"]["bias"] in {"BULLISH", "BEARISH", "SIDEWAYS"}
    assert "ivSummary" in result and "topActivity" in result and "oiAnalysis" in result


def test_options_analytics_endpoint_supports_inst_alias():
    client = TestClient(app)
    response = client.get("/api/options/analytics", params={"inst": "NIFTY", "spot_price": 24530, "iv_pct": 14, "time_to_expiry_days": 2})
    assert response.status_code == 200
    payload = response.json()
    assert payload["index"] == "NIFTY"
    assert payload["atmStrike"] % 50 == 0
    assert payload["dataSource"].startswith("NIFTY")


def test_options_greeks_endpoint():
    client = TestClient(app)
    response = client.get("/api/options/greeks", params={"inst": "SENSEX", "spot_price": 75000, "strike": 75000, "type": "CE"})
    assert response.status_code == 200
    payload = response.json()
    assert payload["type"] == "CE"
    assert payload["strike"] == 75000
    assert 0 < payload["greeks"]["delta"] < 1


def test_options_oi_and_iv_endpoints():
    client = TestClient(app)
    oi_response = client.get("/api/options/oi-analysis", params={"inst": "SENSEX", "spot_price": 75000})
    iv_response = client.get("/api/options/iv-analysis", params={"inst": "SENSEX", "spot_price": 75000, "iv_pct": 16})
    top_activity_response = client.get("/api/options/top-activity", params={"inst": "SENSEX", "spot_price": 75000})
    assert oi_response.status_code == 200
    assert iv_response.status_code == 200
    assert top_activity_response.status_code == 200
    oi_payload = oi_response.json()
    iv_payload = iv_response.json()
    top_activity_payload = top_activity_response.json()
    assert "callLongBuildup" in oi_payload
    assert oi_payload["interpretation"]["bias"] in {"BULLISH", "BEARISH", "NEUTRAL"}
    assert iv_payload["overallIV"] >= 16
    assert 10 <= iv_payload["ivPercentile"] <= 90
    assert len(top_activity_payload["highestVolume"]) == 5


def test_nifty_options_analytics_route():
    client = TestClient(app)
    response = client.get("/api/nifty/options/analytics", params={"spot_price": 24530})
    assert response.status_code == 200
    payload = response.json()
    assert payload["index"] == "NIFTY"
    assert payload["atmStrike"] % 50 == 0


def strategy_context():
    from backend.backtesting.strategies.base import StrategyContext

    return StrategyContext(timeframe="5m", top_signals_per_day=2)
